import boto3
import openpyxl
import io
import json
import os

s3_client = boto3.client('s3')

def lambda_handler(event, context):
    """
    This handler is designed to run within a Step Function.
    It receives the original SOP filename from the previous step,
    finds the corresponding DE template, processes it, and returns
    the S3 location of its JSON output.
    """
    try:
        # Step 1: Get the original SOP filename passed from the previous step.
        # The Step Function passes the entire output from the first Lambda.
        sop_filename = event["sop_filename"]
        if not sop_filename:
            raise ValueError("Input event did not contain a 'sop_filename' key.")
        
        # Step 2: Construct the S3 key for the corresponding DE Template Excel file.
        # e.g., "My SOP File.pdf" becomes "My SOP File.xlsx"
        base_filename = os.path.splitext(sop_filename)[0]
        template_filename = f"{base_filename}.xlsx"
        
        input_bucket = "incoming-sop" # The bucket where DE templates are stored
        input_key = f"DE_Templates/{template_filename}"
        
        print(f"Attempting to process template for '{sop_filename}': s3://{input_bucket}/{input_key}")

        # Step 3: Read the Excel file from S3.
        response = s3_client.get_object(Bucket=input_bucket, Key=input_key)
        file_stream = io.BytesIO(response['Body'].read())
        wb = openpyxl.load_workbook(file_stream, data_only=True)

        if "DE Template" not in wb.sheetnames:
            raise ValueError("Sheet named 'DE Template' not found in the workbook.")

        sheet = wb["DE Template"]

        # Step 4: Locate the header row (your original logic).
        expected_headers = ["Attribute", "Required Questions", "Considerations"]
        header_row_idx = None
        headers = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row is None: continue
            cleaned_row = [str(cell).strip() if cell else "" for cell in row]
            if all(header in cleaned_row for header in expected_headers):
                header_row_idx = i
                headers = cleaned_row
                break
        
        if header_row_idx is None:
            raise ValueError("Required headers ('Attribute', 'Required Questions', 'Considerations') not found.")

        # Step 5: Extract data under the headers (your original logic).
        data = []
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row): continue
            row_dict = dict(zip(headers, row))
            extracted_row = {
                "Attribute": str(row_dict.get("Attribute", "")).strip(),
                "Required Questions": str(row_dict.get("Required Questions", "")).strip(),
                "Considerations": str(row_dict.get("Considerations", "")).strip()
            }
            if extracted_row["Attribute"]:
                data.append(extracted_row)

        # Step 6: Prepare the output path and save the JSON result to S3.
        output_bucket = "de-processing-bucket"
        # The output JSON will have the same base name as the SOP/template.
        output_key = f"DE_Templates/{base_filename}.json"

        s3_client.put_object(
            Bucket=output_bucket,
            Key=output_key,
            Body=json.dumps(data, indent=2),
            ContentType="application/json"
        )
        print(f"Successfully extracted template data to: s3://{output_bucket}/{output_key}")

        # Step 7: Return ONLY the location of the output file.
        # This small payload is passed to the next step.
        return {
            "s3_bucket": output_bucket,
            "s3_key": output_key
        }

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        # Re-raise the exception to make the Step Function task fail correctly.
        raise e

